import re, pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings; warnings.filterwarnings("ignore")

SRC  = "/home/claude/output/02_excel_basic_clean_with_filters.xlsx"
OUT3 = "/home/claude/output/03_python_deep_clean.xlsx"
OUT4 = "/home/claude/output/04_binary_analysis.xlsx"

df = pd.read_excel(SRC)
df.columns = [
    "timestamp","age_group","gender","region","marital_status","occupation",
    "education_level","primary_platform","daily_watch_hours","content_type",
    "peak_usage_time","daily_opens","voice_msg_behavior","usage_duration_since",
    "content_relevance","difficulty_closing_app","productivity_impact",
    "sleep_impact","feeling_after_closing","watching_companion",
    "behavior_while_watching","phone_during_family","family_opinion",
    "reason_for_watching","social_media_without_reels","purchased_from_video",
    "purchase_reason","purchase_influence_level","rewatched_before_purchase",
]

_MD  = re.compile(r"^\*\s*|\*\*(.+?)\*\*")
_EMJ = re.compile(r"[\U0001F300-\U0001FAFF\u2600-\u27BF\u200d]+", re.UNICODE)
def clean_text(v):
    if not isinstance(v, str): return v
    v = _MD.sub(lambda m: m.group(1) or "", v)
    v = _EMJ.sub("", v).strip()
    return re.sub(r"\s{2,}", " ", v).strip()
for c in df.select_dtypes("object").columns: df[c] = df[c].apply(clean_text)

def pick(val, rules):
    if not isinstance(val, str): return val
    for k, r in rules.items():
        if k in val: return r
    return val

def pick_multi(val, rules, sep=" | "):
    if not isinstance(val, str): return val
    out = []
    for t in re.split(r"[;,،]", val):
        t = t.strip()
        matched = next((r for k, r in rules.items() if k in t), None)
        if matched and matched not in out: out.append(matched)
        elif t and not re.search(r'[\u0600-\u06FF]', t) and t not in out: out.append(t)
    return sep.join(out) if out else val

def make_binary(series, prefix):
    all_vals = sorted({item.strip() for v in series.dropna() for item in v.split(" | ") if item.strip()})
    result = {}
    for val in all_vals:
        col = prefix + val.replace(" ","_").replace("&","and").replace("/","_").replace("-","_")
        result[col] = series.apply(lambda x, v=val: 1 if isinstance(x,str) and v in x.split(" | ") else 0)
    return pd.DataFrame(result)

df["age_group"]     = df["age_group"].apply(lambda v: pick(v,{"جيل الجامعة":"18-24","الشغالين":"25-34","أصحاب الخبرة":"35-44","تحت الـ 18":"Under 18","Pros":"45-54","الكبار":"55+","18-24":"18-24","25-34":"25-34","35-44":"35-44","45-54":"45-54","55+":"55+"}))
df["gender"]        = df["gender"].apply(lambda v: pick(v,{"رجل":"Male","شاب":"Male","بنت":"Female","سيدة":"Female"}))
df["region"]        = df["region"].apply(lambda v: pick(v,{"القاهرة الكبرى":"Greater Cairo","خارج مصر":"Abroad","الدلتا":"Nile Delta","الصعيد":"Upper Egypt","الساحل":"North Coast / Red Sea","القناة":"Canal Zone / Sinai"}))
df["marital_status"]= df["marital_status"].apply(lambda v: pick(v,{"لأ لسه":"Single","لا لسه":"Single","وعندي عيال":"Married with Children","مفيش عيال":"Married no Children","مطلق":"Divorced / Widowed","أرمل":"Divorced / Widowed"}))
df["occupation"]    = df["occupation"].apply(lambda v: pick(v,{"موظف":"Employee (Gov/Private)","بدرس":"Student","Freelancer":"Freelancer","عمل حر":"Freelancer","البيت":"Homemaker / Unemployed","متقاعد":"Retired"}))
df["education_level"]= df["education_level"].apply(lambda v: pick(v,{"خريج":"University Graduate","طالب جامعي":"University Student","دراسات عليا":"Postgraduate","قبل الثانوي":"Pre-Secondary Student","متوسط":"Intermediate/Technical Education","فني":"Intermediate/Technical Education","ثانوي":"Secondary School Student","بفك الخط":"Basic Literacy"}))

VALID_PLAT = {"TikTok":"TikTok","Instagram Reels":"Instagram Reels","YouTube Shorts":"YouTube Shorts","Facebook Reels":"Facebook Reels","Snapchat Spotlight":"Snapchat Spotlight","Snapchat":"Snapchat Spotlight"}
def normalize_platform(val):
    if not isinstance(val, str): return "Other"
    result, has_other = [], False
    for t in re.split(r"[;,،|]", val):
        t = t.strip()
        matched = next((en for k,en in VALID_PLAT.items() if k in t), None)
        if matched:
            if matched not in result: result.append(matched)
        elif t: has_other = True
    if has_other: result.append("Other")
    return " | ".join(result) if result else "Other"
df["primary_platform"] = df["primary_platform"].apply(normalize_platform)

df["daily_watch_hours"]= df["daily_watch_hours"].apply(lambda v: pick(v,{"أقل من 30":"< 30 min","30":"30-60 min","أكثر من 3":"3+ hrs","2":"2-3 hrs","1":"1-2 hrs"}))
df["content_type"]     = df["content_type"].apply(lambda v: pick_multi(v,{"كوميدي":"Comedy & Entertainment","تعليمي":"Educational & Cultural","أخبار":"News & Current Affairs","ديني":"Religious","تطوير ذات":"Self-development","رياضة":"Sports","موسيقى":"Music & Dance","طبخ":"Cooking","موضة":"Fashion & Beauty","ألعاب":"Gaming"}))
df["peak_usage_time"]  = df["peak_usage_time"].apply(lambda v: pick(v,{"قبل النوم":"Before sleep","أول ما أصحى":"First thing in morning","أي وقت":"Any free moment (constantly)","constantly":"Any free moment (constantly)","البريك":"During breaks","بأكل":"While eating","المواصلات":"On commute"}))
df["daily_opens"]      = df["daily_opens"].apply(lambda v: pick(v,{"طول اليوم":"All day (lost count)","Lost count":"All day (lost count)","6":"6-10 times/day","3":"3-5 times/day","مرة":"1-2 times/day","مرتين":"1-2 times/day"}))

def voice(v):
    if not isinstance(v, str): return v
    if any(x in v.lower() for x in ["2×","مرتين","x2","2x"]): return "No patience – 2x speed"
    if "أجل" in v: return "Sometimes postpone"
    if "بسرعته" in v: return "Normal speed – no problem"
    if "جزء منه" in v: return "Partial listen if long"
    return v
df["voice_msg_behavior"] = df["voice_msg_behavior"].apply(voice)

df["usage_duration_since"]      = df["usage_duration_since"].apply(lambda v: pick(v,{"سنة إلى سنتين":"1-2 years","سنتين إلى 4":"2-4 years","أكثر من 4":"4+ years","أقل من 6":"< 6 months","6 شهور":"6-12 months"}))
df["content_relevance"]         = df["content_relevance"].apply(lambda v: pick(v,{"مناسبة جدًا":"Very relevant","أحيانًا مناسبة":"Sometimes relevant","مناسبة غالبًا":"Mostly relevant","مش مناسبة":"Not relevant at all"}))
df["difficulty_closing_app"]    = df["difficulty_closing_app"].apply(lambda v: pick(v,{"سهل جدًا":"Very easy to stop","أحيانًا صعب":"Sometimes hard to stop","صعب غالبًا":"Usually hard to stop","صعب جدًا":"Very hard to stop"}))
df["productivity_impact"]       = df["productivity_impact"].apply(lambda v: pick(v,{"مش مأثرة":"No impact","بضيع وقت":"Wastes productive time","بتشت":"Little distracted","بريك حلو":"Good break – helps refocus"}))
df["sleep_impact"]              = df["sleep_impact"].apply(lambda v: pick(v,{"بسهر شوية زيادة":"Sleep slightly later (normal)","نومي بقى وحش":"Disaster (no sleep)","disaster":"Disaster (no sleep)","بنام زي الفل":"No impact on sleep","بنام زى الفل":"No impact on sleep","بسهر لحد الفجر":"Sleep very late (regret it)"}))
df["feeling_after_closing"]     = df["feeling_after_closing"].apply(lambda v: pick(v,{"مبسوط ومنتعش":"Happy and refreshed","ندمان":"Regret wasted time","عادي":"Neutral (nothing special)"}))
df["watching_companion"]        = df["watching_companion"].apply(lambda v: pick(v,{"لوحدي":"Alone (guilty pleasure)","شريك الحياة":"With life partner","الأهل":"With family","الصحاب":"With friends","الشغل":"With classmates/coworkers","الجامعة":"With classmates/coworkers"}))
df["behavior_while_watching"]   = df["behavior_while_watching"].apply(lambda v: pick(v,{"أشاهد فقط":"Watch only","أشارك":"Share videos","أبحث":"Search for more on topic","أحفظ":"Save videos"}))
df["phone_during_family"]       = df["phone_during_family"].apply(lambda v: pick(v,{"بحترم":"No – respect family time","مش قادر":"Yes, mostly (can't put it down)","بشارك الأهل":"Share with family","بشوف بسرعة":"Quick check then back"}))
df["family_opinion"]            = df["family_opinion"].apply(lambda v: pick(v,{"بيشجعوني":"Supportive (edutainment)","بيشتكوا":"Complain – causes conflicts","بينصحوني":"Advise me to reduce usage","مش مهتمين":"Indifferent"}))
df["reason_for_watching"]       = df["reason_for_watching"].apply(lambda v: pick(v,{"تحسين المزاج":"Mood improvement","قتل الوقت":"Killing time","الترفيه":"Entertainment","التعلم":"Learning","الترند":"Following trends"}))
df["social_media_without_reels"]= df["social_media_without_reels"].apply(lambda v: pick(v,{"أقل بكثير":"Would decrease a lot","أقل قليلاً":"Would decrease slightly","لن يتغير":"Would not change","سيزيد بمعدل كبير":"Would increase a lot","ربما يزيد":"Might increase slightly"}))
df["purchased_from_video"]      = df["purchased_from_video"].apply(lambda v: pick(v,{"لأ خالص":"Never","لا خالص":"Never","آه كتير":"Yes, often (easily influenced)","بتعجبني حاجات":"Like items but never buy","مرة أو اتنين":"Once or twice"}))
df["purchase_reason"]           = df["purchase_reason"].apply(lambda v: pick(v,{"لفتت انتباهي":"Visual Attraction","مقنع":"Content Persuasiveness","محتاجها بالفعل":"Need","ترند":"Trend / Social Proof","مشترتش":"N/A – Never purchased"}))
df["purchase_influence_level"]  = df["purchase_influence_level"].apply(lambda v: pick(v,{"أفكر شوية":"Moderate influence","Moderate":"Moderate influence","أقل تأثير":"Low influence","Low":"Low influence","بدون تفكير":"High influence (impulse)","Impulse":"High influence (impulse)"}))
df["rewatched_before_purchase"] = df["rewatched_before_purchase"].apply(lambda v: pick(v,{"غالبًا":"Usually","أبدًا":"Never","أحيانًا":"Sometimes","دايمًا":"Always","نادرًا":"Rarely"}))

for col in df.columns:
    n = df[col].isnull().sum()
    if n > 0: df[col] = df[col].fillna(df[col].mode()[0])

opens_s = {"1-2 times/day":1,"3-5 times/day":2,"6-10 times/day":3,"All day (lost count)":4}
hours_s = {"< 30 min":1,"30-60 min":2,"1-2 hrs":3,"2-3 hrs":4,"3+ hrs":5}
df["watch_intensity_score"] = df.apply(lambda r: int(round((opens_s.get(r["daily_opens"],2)+hours_s.get(r["daily_watch_hours"],3))/2)), axis=1).clip(1,5)
df["user_segment"]          = df["watch_intensity_score"].map({1:"Light User",2:"Light User",3:"Moderate User",4:"Heavy User",5:"Heavy User"})
df["content_count"]         = df["content_type"].apply(lambda v: len(v.split(" | ")) if isinstance(v,str) else 0)
df["is_egypt_resident"]     = (df["region"] != "Abroad").astype(int)

TH=Side(style="thin",color="BFBFBF"); BD=Border(left=TH,right=TH,top=TH,bottom=TH)
HG=PatternFill("solid",fgColor="145A32"); HB=PatternFill("solid",fgColor="1F4E79")
HO=PatternFill("solid",fgColor="7B241C"); HP=PatternFill("solid",fgColor="154360")
AG=PatternFill("solid",fgColor="D5F5E3"); AO=PatternFill("solid",fgColor="FADBD8")
AP=PatternFill("solid",fgColor="D6EAF8"); WF=PatternFill("solid",fgColor="FFFFFF")
HFT=Font(bold=True,color="FFFFFF",name="Arial",size=10); DFT=Font(name="Arial",size=9)
CA=Alignment(horizontal="center",vertical="center",wrap_text=True)
LA=Alignment(horizontal="left",vertical="center",wrap_text=True)
DERIVED={"watch_intensity_score","user_segment","content_count","is_egypt_resident"}

def apply_style(ws, plt_start=None, cnt_start=None):
    for j,cell in enumerate(ws[1],1):
        cn=str(cell.value or "")
        if cnt_start and j>=cnt_start: cell.fill=HP
        elif plt_start and j>=plt_start: cell.fill=HO
        elif cn in DERIVED: cell.fill=HB
        else: cell.fill=HG
        cell.font=HFT; cell.alignment=CA; cell.border=BD
    for i,row in enumerate(ws.iter_rows(min_row=2),2):
        for j,cell in enumerate(row,1):
            if cnt_start and j>=cnt_start: cell.fill=AP if i%2==0 else WF
            elif plt_start and j>=plt_start: cell.fill=AO if i%2==0 else WF
            else: cell.fill=AG if i%2==0 else WF
            cell.font=DFT; cell.alignment=LA; cell.border=BD
    for col in ws.columns:
        mx=max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width=min(mx+3,40)
    ws.freeze_panes="B2"; ws.row_dimensions[1].height=32
    ws.auto_filter.ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

def make_summary(wb,name,rows):
    ws=wb.create_sheet(name)
    for r,rd in enumerate(rows,1):
        for c,v in enumerate(rd,1):
            cell=ws.cell(r,c,v); cell.border=BD; cell.alignment=CA
            if r==1: cell.fill=HG; cell.font=HFT
            elif r==len(rows): cell.fill=AG; cell.font=Font(bold=True,name="Arial",size=10)
            else: cell.fill=AG if r%2==0 else WF; cell.font=DFT
    ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=70

shape3=df.shape; abroad=int((df["region"]=="Abroad").sum())
df.to_excel(OUT3,index=False)
wb3=load_workbook(OUT3); ws3=wb3.active; ws3.title="Deep Clean (EN)"
apply_style(ws3)
make_summary(wb3,"Summary",[["Item","Value"],["Source","File 02 — Excel Basic Clean"],["Type","Wide Format — multi-label ' | '"],["Rows",f"{shape3[0]:,}"],["Columns",f"{shape3[1]}"],["Nulls","0"],["Errors Fixed","9"],["Egypt Residents",f"{shape3[0]-abroad:,}"],["Abroad (flagged)",str(abroad)]])
wb3.save(OUT3)

df04=df.copy()
plt_d=make_binary(df04["primary_platform"],"plt_"); cnt_d=make_binary(df04["content_type"],"cnt_")
df04=df04.drop(columns=["primary_platform","content_type"])
df04=pd.concat([df04,plt_d,cnt_d],axis=1)
n_base=df04.shape[1]-len(plt_d.columns)-len(cnt_d.columns)
plt_start=n_base+1; cnt_start=n_base+len(plt_d.columns)+1
shape04=df04.shape
df04.to_excel(OUT4,index=False)
wb4=load_workbook(OUT4); ws4=wb4.active; ws4.title="Binary Analysis"
apply_style(ws4,plt_start=plt_start,cnt_start=cnt_start)
make_summary(wb4,"Summary",[["Item","Value"],["Source","File 03 — Wide Format"],["Type","Binary 0/1"],["Rows",f"{shape04[0]:,}"],["Columns",str(shape04[1])],["plt_* cols",str(len(plt_d.columns))],["cnt_* cols",str(len(cnt_d.columns))],["SQL Example","WHERE plt_TikTok=1 AND cnt_Religious=1"]])

ws4l=wb4.create_sheet("Column Legend")
legend=[["Column","Meaning","Type"],["plt_TikTok","1=TikTok","Platform"],["plt_Instagram_Reels","1=Instagram Reels","Platform"],["plt_YouTube_Shorts","1=YouTube Shorts","Platform"],["plt_Facebook_Reels","1=Facebook Reels","Platform"],["plt_Snapchat_Spotlight","1=Snapchat Spotlight","Platform"],["plt_Other","1=Other platform","Platform"],["cnt_Comedy_and_Entertainment","1=Comedy & Ent.","Content"],["cnt_Educational_and_Cultural","1=Educational","Content"],["cnt_News_and_Current_Affairs","1=News","Content"],["cnt_Religious","1=Religious","Content"],["cnt_Self_development","1=Self-dev","Content"],["cnt_Sports","1=Sports","Content"],["cnt_Music_and_Dance","1=Music","Content"],["cnt_Cooking","1=Cooking","Content"],["cnt_Fashion_and_Beauty","1=Fashion","Content"],["cnt_Gaming","1=Gaming","Content"],["watch_intensity_score","1-5 scale","Derived"],["user_segment","Light/Moderate/Heavy","Derived"],["content_count","# content types","Derived"],["is_egypt_resident","1=Egypt/0=Abroad","Derived"]]
FM={"Platform":(HO,AO),"Content":(HP,AP),"Derived":(HB,PatternFill("solid",fgColor="D6EAF8"))}
for r,rd in enumerate(legend,1):
    tp=rd[2] if len(rd)>2 else ""
    hf,df_=FM.get(tp,(HG,AG))
    for c,v in enumerate(rd,1):
        cell=ws4l.cell(r,c,v); cell.border=BD
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        if r==1: cell.fill=HG; cell.font=HFT
        else:
            cell.fill=hf if c==1 else df_
            cell.font=Font(bold=(c==1),name="Courier New" if c==1 else "Arial",size=9,color="FFFFFF" if c==1 else "222222")
ws4l.column_dimensions["A"].width=30; ws4l.column_dimensions["B"].width=30; ws4l.column_dimensions["C"].width=14
wb4._sheets=[wb4["Binary Analysis"],wb4["Summary"],wb4["Column Legend"]]
wb4.save(OUT4)
print(f"File 03: {shape3[0]:,}x{shape3[1]}  |  File 04: {shape04[0]:,}x{shape04[1]}")
