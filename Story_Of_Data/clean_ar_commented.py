"""
╔══════════════════════════════════════════════════════════════════════╗
║  تنظيف عميق — من عربي خام (ملف 02) إلى إنجليزي نظيف               ║
║  كود واحد ينتج ملفين:                                               ║
║    03_python_deep_clean.xlsx  → Wide Format   (3289 × 33)           ║
║    04_binary_analysis.xlsx    → Binary 0/1    (3289 × 47)           ║
╚══════════════════════════════════════════════════════════════════════╝

الكود ده بيعمل إيه؟
━━━━━━━━━━━━━━━━━━
§1  تسمية الأعمدة بالإنجليزي + حذف الإيميل
§2  تنظيف النصوص (Markdown + Emoji + whitespace)
§3  دوال الترجمة المساعدة
§4  ترجمة 29 عمود من عربي لإنجليزي + تصحيح 9 أخطاء
§5  ملء الـ Nulls
§6  فحص نهائي
§7  الأعمدة المشتقة (4 أعمدة جديدة)
§8  Validation
§9  حفظ ملف 03 (Wide Format)
§10 إنتاج ملف 04 (Binary 0/1) من نفس الكود
"""

import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ── مسارات الملفات ──────────────────────────────────────────────────
# SRC: نبدأ من ملف 02 (بعد التنظيف الخفيف بالإكسيل)
SRC  = "/home/claude/output/02_excel_basic_clean_with_filters.xlsx"
OUT3 = "/home/claude/output/03_python_deep_clean.xlsx"
OUT4 = "/home/claude/output/04_binary_analysis.xlsx"

# pd.read_excel: بيقرأ ملف xlsx ويحوله لـ DataFrame (جدول بيانات في الذاكرة)
df = pd.read_excel(SRC)
print(f"[Loaded] {df.shape[0]:,} rows × {df.shape[1]} columns")

# ══════════════════════════════════════════════════════════════════════
# § 1 — تسمية الأعمدة بالإنجليزي + حذف الإيميل
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ليه: الأعمدة العربية مش متوافقة مع أدوات التحليل
# إزاي: قائمة ثابتة بالترتيب نفسه وبنعيّنها مرة واحدة
# ══════════════════════════════════════════════════════════════════════
df.columns = [
    "timestamp","age_group","gender","region","marital_status","occupation",
    "education_level","primary_platform","daily_watch_hours","content_type",
    "peak_usage_time","daily_opens","voice_msg_behavior","usage_duration_since",
    "content_relevance","difficulty_closing_app","productivity_impact",
    "sleep_impact","feeling_after_closing","watching_companion",
    "behavior_while_watching","phone_during_family","family_opinion",
    "reason_for_watching","social_media_without_reels","purchased_from_video",
    "purchase_reason","purchase_influence_level","rewatched_before_purchase",
]
print("[§1] Columns renamed ✓")

# ══════════════════════════════════════════════════════════════════════
# § 2 — تنظيف النصوص: Markdown + Emoji + Whitespace
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# خطأ #1: Google Forms حط * **نص** قبل بعض الإجابات
# خطأ #3: Emoji في أول بعض الإجابات بيكسر الترجمة
# re.compile: بيحول النمط لـ regex جاهز للتطبيق السريع
# re.UNICODE: علشان يتعرف على الـ emoji unicode
# ══════════════════════════════════════════════════════════════════════
_MD  = re.compile(r"^\*\s*|\*\*(.+?)\*\*")
_EMJ = re.compile(r"[\U0001F300-\U0001FAFF\u2600-\u27BF\u200d]+", re.UNICODE)

def clean_text(v):
    """
    بتنظف خلية نصية من 3 حاجات:
    1. Markdown: * **نص** → نص
    2. Emoji: 🎓📚🔥 → يشيلهم
    3. مسافات زيادة → مسافة واحدة
    """
    if not isinstance(v, str): return v          # لو مش نص → سيبه
    v = _MD.sub(lambda m: m.group(1) or "", v)   # شيل ** وسيب النص الجوّه
    v = _EMJ.sub("", v).strip()                   # شيل الـ emoji
    return re.sub(r"\s{2,}", " ", v).strip()       # ضغط مسافات زيادة

# تطبيق على كل الأعمدة النصية
# select_dtypes("object"): بيجيب الأعمدة اللي نوعها نص
for c in df.select_dtypes("object").columns:
    df[c] = df[c].apply(clean_text)
print("[§2] Markdown + Emoji removed ✓")

# ══════════════════════════════════════════════════════════════════════
# § 3 — دوال مساعدة للترجمة
# ━━━━━━━━━━━━━━━━━━━━━━━━━━
# pick(): للقيم البسيطة — بتبحث عن أول مفتاح موجود في النص وترجعه
# pick_multi(): للقيم المركبة المفصولة بـ ; أو , — بتترجم كل token
# make_binary(): بتحول ' | ' column لأعمدة 0/1 (مش get_dummies العادية
#   لأن get_dummies بتقسم على المسافة وبتكسر 'Facebook Reels' لكلمتين)
# ══════════════════════════════════════════════════════════════════════
def pick(val, rules):
    """بترجع أول ترجمة يتطابق مفتاحها مع جزء من النص"""
    if not isinstance(val, str): return val
    for k, r in rules.items():
        if k in val: return r
    return val

def pick_multi(val, rules, sep=" | "):
    """
    بتترجم قيم مركبة (مفصولة بـ ; أو ,)
    الـ separator الموحد في المخرج: ' | ' (pipe)
    ' | ' هو industry standard في multi-label data
    بيتفهم مباشرة في Power BI / SQL Server / Python
    """
    if not isinstance(val, str): return val
    out = []
    for t in re.split(r"[;,،]", val):
        t = t.strip()
        matched = next((r for k, r in rules.items() if k in t), None)
        if matched and matched not in out: out.append(matched)
        elif t and not re.search(r'[\u0600-\u06FF]', t) and t not in out: out.append(t)
    return sep.join(out) if out else val

def make_binary(series, prefix):
    """
    تحويل multi-label column (بـ ' | ') لأعمدة binary 0/1
    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ليه مش get_dummies؟
      get_dummies بتقسم على أي مسافة:
      'Facebook Reels' → 'Facebook'=1, 'Reels'=1   ← غلط
    ليه make_binary؟
      بتقسم على ' | ' الكامل كـ delimiter:
      'Facebook Reels | TikTok' → plt_Facebook_Reels=1, plt_TikTok=1   ← صح
    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    مناسب لـ SQL Server: WHERE plt_TikTok=1 AND cnt_Religious=1
    """
    # جمع كل القيم الفريدة من كل الخلايا
    all_vals = sorted({
        item.strip()
        for v in series.dropna()
        for item in v.split(" | ")
        if item.strip()
    })
    result = {}
    for val in all_vals:
        # تحويل اسم القيمة لاسم عمود صالح
        col_name = (prefix
                    + val.replace(" ","_")
                        .replace("&","and")
                        .replace("/","_")
                        .replace("-","_"))
        # 1 لو القيمة موجودة في الخلية، 0 لو لأ
        result[col_name] = series.apply(
            lambda x, v=val: 1 if isinstance(x,str) and v in x.split(" | ") else 0)
    return pd.DataFrame(result)

# ══════════════════════════════════════════════════════════════════════
# § 4 — ترجمة 29 عمود + تصحيح 9 أخطاء
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# كل عمود ليه قاموس rules خاص بيه
# pick/pick_multi بيمشوا عليه ويترجموا
# الأخطاء مكتوبة جنب كل عمود متأثر
# ══════════════════════════════════════════════════════════════════════

# age_group
df["age_group"] = df["age_group"].apply(lambda v: pick(v, {
    "جيل الجامعة":"18-24","الشغالين":"25-34","أصحاب الخبرة":"35-44",
    "تحت الـ 18":"Under 18","Pros":"45-54","الكبار":"55+",
    "18-24":"18-24","25-34":"25-34","35-44":"35-44","45-54":"45-54","55+":"55+",
}))

# gender
df["gender"] = df["gender"].apply(lambda v: pick(v,{
    "رجل":"Male","شاب":"Male","بنت":"Female","سيدة":"Female",
}))

# region | خطأ #5: Abroad (مش Outside Egypt)
df["region"] = df["region"].apply(lambda v: pick(v,{
    "القاهرة الكبرى":"Greater Cairo","خارج مصر":"Abroad",
    "الدلتا":"Nile Delta","الصعيد":"Upper Egypt",
    "الساحل":"North Coast / Red Sea","القناة":"Canal Zone / Sinai",
}))

# marital_status
df["marital_status"] = df["marital_status"].apply(lambda v: pick(v,{
    "لأ لسه":"Single","لا لسه":"Single",
    "وعندي عيال":"Married with Children","مفيش عيال":"Married no Children",
    "مطلق":"Divorced / Widowed","أرمل":"Divorced / Widowed",
}))

# occupation
df["occupation"] = df["occupation"].apply(lambda v: pick(v,{
    "موظف":"Employee (Gov/Private)","بدرس":"Student",
    "Freelancer":"Freelancer","عمل حر":"Freelancer",
    "البيت":"Homemaker / Unemployed","متقاعد":"Retired",
}))

# education_level | خطأ #1: Intermediate/Technical (مش Vocational)
df["education_level"] = df["education_level"].apply(lambda v: pick(v,{
    "خريج":"University Graduate","طالب جامعي":"University Student",
    "دراسات عليا":"Postgraduate","قبل الثانوي":"Pre-Secondary Student",
    "متوسط":"Intermediate/Technical Education",
    "فني":"Intermediate/Technical Education",
    "ثانوي":"Secondary School Student","بفك الخط":"Basic Literacy",
}))

# primary_platform — 5 منصات معتمدة + Other لأي حاجة تانية
# normalize_platform: بتمر على كل منصة في الخلية
# لو من الـ 5 المعتمدة → تفضل | لو غيرهم → Other
VALID_PLAT = {
    "TikTok":"TikTok","Instagram Reels":"Instagram Reels",
    "YouTube Shorts":"YouTube Shorts","Facebook Reels":"Facebook Reels",
    "Snapchat Spotlight":"Snapchat Spotlight","Snapchat":"Snapchat Spotlight",
}
def normalize_platform(val):
    if not isinstance(val, str): return "Other"
    result, has_other = [], False
    for t in re.split(r"[;,،|]", val):
        t = t.strip()
        matched = next((en for k,en in VALID_PLAT.items() if k in t), None)
        if matched:
            if matched not in result: result.append(matched)
        elif t: has_other = True
    if has_other: result.append("Other")
    return " | ".join(result) if result else "Other"
df["primary_platform"] = df["primary_platform"].apply(normalize_platform)

# daily_watch_hours
df["daily_watch_hours"] = df["daily_watch_hours"].apply(lambda v: pick(v,{
    "أقل من 30":"< 30 min","30":"30-60 min",
    "أكثر من 3":"3+ hrs","2":"2-3 hrs","1":"1-2 hrs",
}))

# content_type — multi-label بـ ' | ' (قرار واعي: يفضل محتواه)
CONTENT_MAP = {
    "كوميدي":"Comedy & Entertainment","تعليمي":"Educational & Cultural",
    "أخبار":"News & Current Affairs","ديني":"Religious",
    "تطوير ذات":"Self-development","رياضة":"Sports",
    "موسيقى":"Music & Dance","طبخ":"Cooking",
    "موضة":"Fashion & Beauty","ألعاب":"Gaming",
}
df["content_type"] = df["content_type"].apply(
    lambda v: pick_multi(v, CONTENT_MAP, sep=" | "))

# peak_usage_time
df["peak_usage_time"] = df["peak_usage_time"].apply(lambda v: pick(v,{
    "قبل النوم":"Before sleep","أول ما أصحى":"First thing in morning",
    "أي وقت":"Any free moment (constantly)","constantly":"Any free moment (constantly)",
    "البريك":"During breaks","بأكل":"While eating","المواصلات":"On commute",
}))

# daily_opens | خطأ #4: Arabic leftover في بعض الخلايا
df["daily_opens"] = df["daily_opens"].apply(lambda v: pick(v,{
    "طول اليوم":"All day (lost count)","Lost count":"All day (lost count)",
    "6":"6-10 times/day","3":"3-5 times/day",
    "مرة":"1-2 times/day","مرتين":"1-2 times/day",
}))

# voice_msg_behavior | خطأ #2: 3 تهجئات عربية مختلفة لـ 2x speed → توحيد
# any(): بتشيك لو أي keyword من القائمة موجود في النص
def voice(v):
    if not isinstance(v, str): return v
    if any(x in v.lower() for x in ["2×","مرتين","x2","2x"]): return "No patience – 2x speed"
    if "أجل" in v: return "Sometimes postpone"
    if "بسرعته" in v: return "Normal speed – no problem"
    if "جزء منه" in v: return "Partial listen if long"
    return v
df["voice_msg_behavior"] = df["voice_msg_behavior"].apply(voice)

# usage_duration_since
df["usage_duration_since"] = df["usage_duration_since"].apply(lambda v: pick(v,{
    "سنة إلى سنتين":"1-2 years","سنتين إلى 4":"2-4 years",
    "أكثر من 4":"4+ years","أقل من 6":"< 6 months","6 شهور":"6-12 months",
}))

# content_relevance | خطأ #9: Very relevant + قوس عربي | خطأ #3: emoji (اتحل §2)
df["content_relevance"] = df["content_relevance"].apply(lambda v: pick(v,{
    "مناسبة جدًا":"Very relevant","أحيانًا مناسبة":"Sometimes relevant",
    "مناسبة غالبًا":"Mostly relevant","مش مناسبة":"Not relevant at all",
}))

# difficulty_closing_app
df["difficulty_closing_app"] = df["difficulty_closing_app"].apply(lambda v: pick(v,{
    "سهل جدًا":"Very easy to stop","أحيانًا صعب":"Sometimes hard to stop",
    "صعب غالبًا":"Usually hard to stop","صعب جدًا":"Very hard to stop",
}))

# productivity_impact | خطأ #6: Little distracted (تهجئتان: بتشتتني + بتشتني)
# "بتشت" بيشمل الاثنين لأنها الـ prefix المشترك
df["productivity_impact"] = df["productivity_impact"].apply(lambda v: pick(v,{
    "مش مأثرة":"No impact","بضيع وقت":"Wastes productive time",
    "بتشت":"Little distracted","بريك حلو":"Good break – helps refocus",
}))

# sleep_impact | خطأ #7: Disaster (no sleep) مش Severely disrupted
df["sleep_impact"] = df["sleep_impact"].apply(lambda v: pick(v,{
    "بسهر شوية زيادة":"Sleep slightly later (normal)",
    "نومي بقى وحش":"Disaster (no sleep)","disaster":"Disaster (no sleep)",
    "بنام زي الفل":"No impact on sleep","بنام زى الفل":"No impact on sleep",
    "بسهر لحد الفجر":"Sleep very late (regret it)",
}))

# feeling_after_closing
df["feeling_after_closing"] = df["feeling_after_closing"].apply(lambda v: pick(v,{
    "مبسوط ومنتعش":"Happy and refreshed",
    "ندمان":"Regret wasted time","عادي":"Neutral (nothing special)",
}))

# watching_companion | خطأ #8: 4 قيم ترجمتهم كانت غلط
df["watching_companion"] = df["watching_companion"].apply(lambda v: pick(v,{
    "لوحدي":"Alone (guilty pleasure)","شريك الحياة":"With life partner",
    "الأهل":"With family","الصحاب":"With friends",
    "الشغل":"With classmates/coworkers","الجامعة":"With classmates/coworkers",
}))

# behavior_while_watching
df["behavior_while_watching"] = df["behavior_while_watching"].apply(lambda v: pick(v,{
    "أشاهد فقط":"Watch only","أشارك":"Share videos",
    "أبحث":"Search for more on topic","أحفظ":"Save videos",
}))

# phone_during_family
df["phone_during_family"] = df["phone_during_family"].apply(lambda v: pick(v,{
    "بحترم":"No – respect family time",
    "مش قادر":"Yes, mostly (can't put it down)",
    "بشارك الأهل":"Share with family","بشوف بسرعة":"Quick check then back",
}))

# family_opinion
df["family_opinion"] = df["family_opinion"].apply(lambda v: pick(v,{
    "بيشجعوني":"Supportive (edutainment)","بيشتكوا":"Complain – causes conflicts",
    "بينصحوني":"Advise me to reduce usage","مش مهتمين":"Indifferent",
}))

# reason_for_watching
df["reason_for_watching"] = df["reason_for_watching"].apply(lambda v: pick(v,{
    "تحسين المزاج":"Mood improvement","قتل الوقت":"Killing time",
    "الترفيه":"Entertainment","التعلم":"Learning","الترند":"Following trends",
}))

# social_media_without_reels
df["social_media_without_reels"] = df["social_media_without_reels"].apply(lambda v: pick(v,{
    "أقل بكثير":"Would decrease a lot","أقل قليلاً":"Would decrease slightly",
    "لن يتغير":"Would not change","سيزيد بمعدل كبير":"Would increase a lot",
    "ربما يزيد":"Might increase slightly",
}))

# purchased_from_video
df["purchased_from_video"] = df["purchased_from_video"].apply(lambda v: pick(v,{
    "لأ خالص":"Never","لا خالص":"Never",
    "آه كتير":"Yes, often (easily influenced)",
    "بتعجبني حاجات":"Like items but never buy","مرة أو اتنين":"Once or twice",
}))

# purchase_reason
df["purchase_reason"] = df["purchase_reason"].apply(lambda v: pick(v,{
    "لفتت انتباهي":"Visual Attraction","مقنع":"Content Persuasiveness",
    "محتاجها بالفعل":"Need","ترند":"Trend / Social Proof",
    "مشترتش":"N/A – Never purchased",
}))

# purchase_influence_level
df["purchase_influence_level"] = df["purchase_influence_level"].apply(lambda v: pick(v,{
    "أفكر شوية":"Moderate influence","Moderate":"Moderate influence",
    "أقل تأثير":"Low influence","Low":"Low influence",
    "بدون تفكير":"High influence (impulse)","Impulse":"High influence (impulse)",
}))

# rewatched_before_purchase
df["rewatched_before_purchase"] = df["rewatched_before_purchase"].apply(lambda v: pick(v,{
    "غالبًا":"Usually","أبدًا":"Never","أحيانًا":"Sometimes",
    "دايمًا":"Always","نادرًا":"Rarely",
}))

print("[§4] All 29 columns translated ✓  |  9 errors fixed ✓")

# ══════════════════════════════════════════════════════════════════════
# § 5 — ملء الـ Nulls المتبقية
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# fillna(mode()[0]): بيملأ الفراغات بأكثر قيمة متكررة في العمود
# mode() بترجع Series → [0] عشان ناخد أول قيمة بس
# ══════════════════════════════════════════════════════════════════════
for col in df.columns:
    n = df[col].isnull().sum()
    if n > 0:
        df[col] = df[col].fillna(df[col].mode()[0])
        print(f"   [{col}] filled {n} nulls with mode")
print("[§5] Nulls filled ✓")

# ══════════════════════════════════════════════════════════════════════
# § 6 — فحص نهائي: هل في نص عربي متبقي؟
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# re.compile(r'[\u0600-\u06FF]'): نمط يكشف الحروف العربية
# unicode range للعربية: U+0600 → U+06FF
# ══════════════════════════════════════════════════════════════════════
_AR = re.compile(r'[\u0600-\u06FF]')
remaining = {
    c: df[c].apply(lambda v: bool(_AR.search(str(v))) if isinstance(v,str) else False).sum()
    for c in df.select_dtypes("object").columns
}
remaining = {c:n for c,n in remaining.items() if n > 0}
print("[§6] Arabic remaining:", remaining if remaining else "None ✓")

# ══════════════════════════════════════════════════════════════════════
# § 7 — الأعمدة المشتقة (Derived Columns)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# أعمدة جديدة بتتحسب من الأعمدة الموجودة
# مش موجودة في الاستبيان الأصلي — بنستنتجها
# ══════════════════════════════════════════════════════════════════════

# watch_intensity_score: مقياس 1-5 بيمثل شدة الاستخدام
# بيتحسب من: (opens_score + hours_score) / 2  ثم round
opens_s = {"1-2 times/day":1,"3-5 times/day":2,"6-10 times/day":3,"All day (lost count)":4}
hours_s = {"< 30 min":1,"30-60 min":2,"1-2 hrs":3,"2-3 hrs":4,"3+ hrs":5}
df["watch_intensity_score"] = df.apply(
    lambda r: int(round(
        (opens_s.get(r["daily_opens"],2) + hours_s.get(r["daily_watch_hours"],3)) / 2
    )), axis=1).clip(1, 5)
# clip(1,5): بتضمن إن القيمة مش تقل عن 1 ولا تعدى 5

# user_segment: تصنيف المستخدم بناءً على الـ score
# .map(): بتستبدل كل قيمة في العمود بالمقابل في الـ dict
df["user_segment"] = df["watch_intensity_score"].map(
    {1:"Light User",2:"Light User",3:"Moderate User",4:"Heavy User",5:"Heavy User"})

# content_count: عدد أنواع المحتوى اللي اختارها المستخدم
# .split(" | "): بتقسم النص على ' | ' وترجع list
# len(): بتعد عدد العناصر
df["content_count"] = df["content_type"].apply(
    lambda v: len(v.split(" | ")) if isinstance(v,str) else 0)

# is_egypt_resident: flag — 1 لو مقيم مصر، 0 لو خارجها
# .astype(int): بتحول True/False لـ 1/0
df["is_egypt_resident"] = (df["region"] != "Abroad").astype(int)

print("[§7] Derived columns added ✓ (watch_intensity_score, user_segment, content_count, is_egypt_resident)")

# ══════════════════════════════════════════════════════════════════════
# § 8 — Validation: تأكيد إن القيم صح
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# بنقارن القيم الفعلية بالقيم المتوقعة
# لو لقينا قيمة غريبة → بنطبعها كـ warning
# ══════════════════════════════════════════════════════════════════════
EXPECTED = {
    "gender":              {"Male","Female"},
    "sleep_impact":        {"Sleep slightly later (normal)","No impact on sleep",
                            "Disaster (no sleep)","Sleep very late (regret it)"},
    "productivity_impact": {"No impact","Wastes productive time",
                            "Little distracted","Good break – helps refocus"},
    "watching_companion":  {"Alone (guilty pleasure)","With life partner",
                            "With family","With friends","With classmates/coworkers"},
    "content_relevance":   {"Very relevant","Sometimes relevant",
                            "Mostly relevant","Not relevant at all"},
}
ok = True
for col, exp in EXPECTED.items():
    bad = set(df[col].dropna().unique()) - exp
    if bad: print(f"  ⚠️ {col}: unexpected → {bad}"); ok = False
print("[§8] Validation:", "All values correct ✓" if ok else "Check warnings above")

# ══════════════════════════════════════════════════════════════════════
# § 9 — حفظ ملف 03 (Wide Format) + تنسيق
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Wide Format: الداتا كما هي — كل اختيارات المنصة/المحتوى في خلية واحدة
# مناسب لـ: التقارير العامة + Tableau + Power BI (بـ Split Column)
# ══════════════════════════════════════════════════════════════════════
shape3 = df.shape
abroad = int((df["region"] == "Abroad").sum())
print(f"\n[§9] File 03: {shape3[0]:,} rows × {shape3[1]} columns | nulls={df.isnull().sum().sum()}")

df.to_excel(OUT3, index=False)

# ── ستايل helpers (مشتركة بين ملف 03 و04) ────────────────────────
TH  = Side(style="thin", color="BFBFBF")
BD  = Border(left=TH, right=TH, top=TH, bottom=TH)
HG  = PatternFill("solid", fgColor="145A32")   # أخضر = أعمدة أصلية
HB  = PatternFill("solid", fgColor="1F4E79")   # أزرق = مشتقة
HO  = PatternFill("solid", fgColor="7B241C")   # أحمر = platform binary
HP  = PatternFill("solid", fgColor="154360")   # أزرق داكن = content binary
AG  = PatternFill("solid", fgColor="D5F5E3")
AO  = PatternFill("solid", fgColor="FADBD8")
AP  = PatternFill("solid", fgColor="D6EAF8")
WF  = PatternFill("solid", fgColor="FFFFFF")
HFT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DFT = Font(name="Arial", size=9)
CA  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LA  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
DERIVED = {"watch_intensity_score","user_segment","content_count","is_egypt_resident"}

def apply_style(ws, plt_start=None, cnt_start=None, h_fill=HG):
    """تطبيق ستايل موحد على أي شيت"""
    for j, cell in enumerate(ws[1], 1):
        cname = str(cell.value or "")
        if cnt_start and j >= cnt_start:    cell.fill = HP
        elif plt_start and j >= plt_start:  cell.fill = HO
        elif cname in DERIVED:              cell.fill = HB
        else:                               cell.fill = h_fill
        cell.font=HFT; cell.alignment=CA; cell.border=BD
    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
        for j, cell in enumerate(row, 1):
            if cnt_start and j >= cnt_start:   cell.fill=AP if i%2==0 else WF
            elif plt_start and j >= plt_start: cell.fill=AO if i%2==0 else WF
            else:                              cell.fill=AG if i%2==0 else WF
            cell.font=DFT; cell.alignment=LA; cell.border=BD
    for col in ws.columns:
        mx = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(mx+3, 40)
    ws.freeze_panes="B2"; ws.row_dimensions[1].height=32
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

def make_summary_sheet(wb, sheet_name, rows):
    """إنشاء شيت Summary بتنسيق موحد"""
    ws = wb.create_sheet(sheet_name)
    for r, rd in enumerate(rows, 1):
        for c, v in enumerate(rd, 1):
            cell = ws.cell(r, c, v); cell.border=BD; cell.alignment=CA
            if r==1: cell.fill=HG; cell.font=HFT
            elif r==len(rows): cell.fill=AG; cell.font=Font(bold=True,name="Arial",size=10)
            else:
                cell.fill=AG if r%2==0 else WF; cell.font=DFT
    ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=70
    return ws

wb3 = load_workbook(OUT3)
ws3 = wb3.active; ws3.title="Deep Clean (EN)"
apply_style(ws3)

make_summary_sheet(wb3, "Summary", [
    ["Item",              "Value"],
    ["Source",            "File 02 — Excel Basic Clean (Arabic)"],
    ["Type",              "Wide Format — multi-label separated by ' | '"],
    ["Rows",              f"{shape3[0]:,}"],
    ["Columns",           f"{shape3[1]}  (29 original + 4 derived)"],
    ["Nulls",             "0"],
    ["Errors Fixed",      "9 documented errors"],
    ["Separator",         "' | ' (pipe) — compatible with Power BI / SQL Server / Python"],
    ["Platforms",         "TikTok | Instagram Reels | YouTube Shorts | Facebook Reels | Snapchat Spotlight | Other"],
    ["Egypt Residents",   f"{shape3[0]-abroad:,}"],
    ["Abroad (flagged)",  str(abroad)],
    ["Next File",         "04_binary_analysis.xlsx — Binary 0/1 encoding"],
])
wb3.save(OUT3)
print(f"[§9] File 03 saved ✓")

# ══════════════════════════════════════════════════════════════════════
# § 10 — إنتاج ملف 04 (Binary Encoding) من نفس الكود
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ليه ملف منفصل؟
#   الـ binary columns (0/1) بتكبّر عدد الأعمدة كتير
#   وبتغني عن عمودي primary_platform و content_type الأصليين
# ليه مش بنكبّر الصفوف؟
#   explode() كانت هتضاعف الصفوف — ده مش مطلوب
#   make_binary() بتضيف أعمدة بس مش صفوف → 3289 يفضل 3289
# ══════════════════════════════════════════════════════════════════════
df04 = df.copy()

# make_binary بتحول ' | ' column لأعمدة 0/1
# كل منصة = عمود → plt_TikTok, plt_Facebook_Reels, ...
platform_dummies = make_binary(df04["primary_platform"], "plt_")

# كل نوع محتوى = عمود → cnt_Religious, cnt_Cooking, ...
content_dummies  = make_binary(df04["content_type"],     "cnt_")

# حذف الأعمدة الأصلية من ملف 04 (الـ binary بيغني عنهم)
# لو محتاجهم Wide Format → ارجع لملف 03
df04 = df04.drop(columns=["primary_platform","content_type"])

# pd.concat: بتدمج الـ DataFrames جنب بعض (axis=1 = أعمدة)
df04 = pd.concat([df04, platform_dummies, content_dummies], axis=1)

n_base = df04.shape[1] - len(platform_dummies.columns) - len(content_dummies.columns)
n_plt  = len(platform_dummies.columns)
n_cnt  = len(content_dummies.columns)
plt_start = n_base + 1
cnt_start  = n_base + n_plt + 1
shape04 = df04.shape

print(f"\n[§10] File 04: {shape04[0]:,} rows × {shape04[1]} columns")
print(f"      {n_base} base + {n_plt} plt_* + {n_cnt} cnt_* = {shape04[1]}")
print(f"      Platform cols: {list(platform_dummies.columns)}")
print(f"      Content cols:  {list(content_dummies.columns)}")

df04.to_excel(OUT4, index=False)

wb4 = load_workbook(OUT4)
ws4 = wb4.active; ws4.title="Binary Analysis"
apply_style(ws4, plt_start=plt_start, cnt_start=cnt_start)

# Summary ملف 04
make_summary_sheet(wb4, "Summary", [
    ["Item",               "Value"],
    ["Source",             "File 03 — Deep Clean Wide Format"],
    ["Type",               "Binary Multi-label Encoding (0 / 1)"],
    ["Rows",               f"{shape04[0]:,}  (same 3,289 — no expansion)"],
    ["Total Columns",      str(shape04[1])],
    ["Base Columns",       f"{n_base}  (green header)"],
    ["Derived Columns",    "4  (blue header)"],
    ["Platform Binary",    f"{n_plt}  plt_*  (red header)"],
    ["Content Binary",     f"{n_cnt}  cnt_*  (dark blue header)"],
    ["Binary Values",      "1 = selected  /  0 = not selected"],
    ["Dropped",            "primary_platform, content_type → use File 03 for Wide Format"],
    ["SQL Example",        "SELECT * FROM survey WHERE plt_TikTok=1 AND cnt_Religious=1"],
])

# Column Legend ملف 04
ws4l = wb4.create_sheet("Column Legend")
TYPE_FILLS = {
    "Platform":"Platform Binary","Content":"Content Binary","Derived":"Derived"
}
legend = [
    ["Column",                         "Meaning",                                    "Type"],
    ["plt_TikTok",                     "1 = uses TikTok",                            "Platform"],
    ["plt_Instagram_Reels",            "1 = uses Instagram Reels",                   "Platform"],
    ["plt_YouTube_Shorts",             "1 = uses YouTube Shorts",                    "Platform"],
    ["plt_Facebook_Reels",             "1 = uses Facebook Reels",                    "Platform"],
    ["plt_Snapchat_Spotlight",         "1 = uses Snapchat Spotlight",                "Platform"],
    ["plt_Other",                      "1 = uses another platform",                  "Platform"],
    ["cnt_Comedy_and_Entertainment",   "1 = watches Comedy & Entertainment",         "Content"],
    ["cnt_Educational_and_Cultural",   "1 = watches Educational & Cultural",         "Content"],
    ["cnt_News_and_Current_Affairs",   "1 = watches News & Current Affairs",         "Content"],
    ["cnt_Religious",                  "1 = watches Religious content",              "Content"],
    ["cnt_Self_development",           "1 = watches Self-development",               "Content"],
    ["cnt_Sports",                     "1 = watches Sports",                         "Content"],
    ["cnt_Music_and_Dance",            "1 = watches Music & Dance",                  "Content"],
    ["cnt_Cooking",                    "1 = watches Cooking",                        "Content"],
    ["cnt_Fashion_and_Beauty",         "1 = watches Fashion & Beauty",               "Content"],
    ["cnt_Gaming",                     "1 = watches Gaming",                         "Content"],
    ["watch_intensity_score",          "1–5 intensity scale (opens + hours / 2)",    "Derived"],
    ["user_segment",                   "Light / Moderate / Heavy User",              "Derived"],
    ["content_count",                  "Number of content types selected",           "Derived"],
    ["is_egypt_resident",              "1 = Egypt resident / 0 = Abroad",            "Derived"],
]
FILL_MAP = {
    "Platform": (HO, AO), "Content": (HP, AP),
    "Derived":  (HB, PatternFill("solid", fgColor="D6EAF8")),
}
for r, rd in enumerate(legend, 1):
    tp = rd[2] if len(rd)>2 else ""
    h_f, d_f = FILL_MAP.get(tp, (HG, AG))
    for c_idx, v in enumerate(rd, 1):
        cell = ws4l.cell(r, c_idx, v); cell.border=BD
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        if r==1: cell.fill=HG; cell.font=HFT
        else:
            cell.fill = h_f if c_idx==1 else d_f
            cell.font = Font(
                bold=(c_idx==1),
                name="Courier New" if c_idx==1 else "Arial",
                size=9, color="FFFFFF" if c_idx==1 else "222222")
ws4l.column_dimensions["A"].width=30
ws4l.column_dimensions["B"].width=42
ws4l.column_dimensions["C"].width=18

wb4._sheets = [wb4["Binary Analysis"], wb4["Summary"], wb4["Column Legend"]]
wb4.save(OUT4)
print(f"[§10] File 04 saved ✓  Sheets: {wb4.sheetnames}")

print(f"\n{'='*55}")
print(f"  ✅ DONE — One script, two clean files")
print(f"  File 03 Wide   → {shape3[0]:,} × {shape3[1]}")
print(f"  File 04 Binary → {shape04[0]:,} × {shape04[1]}")
print(f"{'='*55}")
